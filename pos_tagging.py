import konlpy
import operator
from konlpy.tag import Kkma
from konlpy.tag import Twitter
from openpyxl import load_workbook
from openpyxl import Workbook
from konlpy.utils import pprint
from collections import Counter


List = ['SK', 'KT', 'LG', '통합']

load_excel = load_workbook(filename='output.xlsx', data_only=True)
load_sheet = load_excel['SK']

twit = Twitter()
posList, posCountList = set(), {}

for idx in range(4):
    nounList, nounCountList = set(), {}
    load_sheet = load_excel[List[idx]]
    if idx == 0:
        k = 744
    elif idx == 1:
        k = 769
    elif idx == 2:
        k = 764
    else:
        k = 2273

    for i in range(3,k):
        print(i-2,'번째 문장입니다.')
        sentence = twit.pos( load_sheet.cell(row=i, column=3).value + ' ' + str(load_sheet.cell(row=i, column=4).value) + ' ' + str(load_sheet.cell(row=i, column=7).value) )
        for j in sentence:
            tmp = '/'.join(j)
            if tmp not in posList:
                posList.add(tmp)
                posCountList[tmp] = 0
            posCountList[tmp] = posCountList[tmp] + 1

    result = sorted(posCountList.items(), key=lambda i : i[1], reverse=True)
    read_excel = load_workbook(filename='result(Twitter)_pos.xlsx', data_only=True)
    read_sheet = read_excel[List[idx]]
    for t, i in enumerate(result):
        read_sheet.cell(row=t+1, column=1, value=i[0])
        read_sheet.cell(row=t+1, column=2, value=i[1])
    read_excel.save(filename='result(Twitter)_pos.xlsx')


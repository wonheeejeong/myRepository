from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook  #엑셀 읽기용
from openpyxl import Workbook       #엑실 쓰기용
import time

class List:
    url = '주소'
    title = '제목'
    contents = '내용'
    comments = '댓글'
    date = '날짜'
    viewCount = 0
    def __init__(self, _url, _title, _contents, _comments, _date, _viewCount):
        self.url = _url
        self.title = _title
        self.contents = _contents
        self.comments = _comments
        self.date = _date
        self.viewCount = _viewCount

load_excel = load_workbook("output.xlsx", data_only=True)
load_sheet = load_excel['Sheet2']

path = "C:/chromedriver.exe"
driver = webdriver.Chrome(path)

driver.get(load_sheet.cell(row=1, column=3).value)


idx, curPageNum = load_sheet.cell(row=1, column=1).value, load_sheet.cell(row=1, column=2).value
isLastPage = False
db = [List('','','','','',0)]
blacklist = ['아우슈리네', '헹복전도사', '초등교육학', '핑핑이', '보트피플', '게임방송', '잡담']
checkBlacklist = False
driver.refresh()

while curPageNum <= 33:

    driver.refresh()
    postLists = driver.find_elements_by_xpath("//ul[@class='sch_result_list']/li/a[@class='tit_txt']")


    for j in postLists:
        driver.implicitly_wait(1)
        for item in blacklist:
            if (j.text).find(item) >= 0:
                checkBlacklist = True
                continue
        if checkBlacklist:
            checkBlacklist = False
            continue

        j.click()
        driver.switch_to.window(driver.window_handles[-1])

        if( (driver.current_url).find('no=') == -1):
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            driver.implicitly_wait(3)
            continue

        idx = idx + 1
        print(idx,'번째 글입니다')
        driver.refresh()
        title = driver.find_element_by_xpath("//h3[@class='title ub-word']/span[@class='title_subject']")
        print(title.text)

        content = driver.find_elements_by_xpath("//div[@style='overflow:hidden;']/p | //div[@style='overflow:hidden;']/div")
        content_tmp = ''
        for k in content:
            if(k.text != ''):
                #print(k.text)
                content_tmp = content_tmp + k.text + ' '

        comment = driver.find_elements_by_xpath("//p[@class='usertxt ub-word']")
        comment_tmp = ''
        for k in comment:
            if(k.text != ''):
                #print(k.text)
                comment_tmp = comment_tmp + k.text + ' '

        date = driver.find_element_by_xpath("//div[@class='fl']/span[@class='gall_date']")
        view = driver.find_element_by_xpath("//div[@class='fr']/span[@class='gall_count']")

        load_sheet.cell(row=idx + 2, column=2, value=idx)
        load_sheet.cell(row=idx + 2, column=3, value=title.text)
        load_sheet.cell(row=idx + 2, column=4, value=content_tmp)
        load_sheet.cell(row=idx + 2, column=7, value=comment_tmp)
        load_sheet.cell(row=idx + 2, column=5, value=date.text[:10])
        load_sheet.cell(row=idx + 2, column=6, value=view.text[3:])

        load_sheet.cell(row=1, column=1, value=idx)
        load_excel.save(filename='output.xlsx')

        #db.append(List(driver.current_url, title.text, content_tmp, comment_tmp, date.text[:10],int(view.text[3:])))

        #print(idx, db[idx].url, db[idx].title, db[idx].date, db[idx].viewCount)
        #print(db[idx].contents)
        #print(db[idx].comments)

        #driver.implicitly_wait(1)
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        #driver.implicitly_wait(1)

    curPageNum = curPageNum + 1
    load_sheet.cell(row=1, column=2, value=curPageNum)

    if curPageNum % 10 == 1:
        isLastPage = True

    if isLastPage:
        driver.refresh()
        next = driver.find_elements_by_xpath("//div[@class='bottom_paging_box']/a[@class='b_next']")
        for i in next:
            if i.text == '다음':
                i.click()
                isLastPage = False
                driver.implicitly_wait(1)
                break

    else:
        pageLists = driver.find_elements_by_xpath("//div[@class='bottom_paging_box']/a")
        for i in pageLists:
            if i.text == str(curPageNum):
                i.click()
                load_sheet.cell(row=1, column=3, value=driver.current_url)
                break
    #driver.implicitly_wait(3)
    driver.refresh()


#write_excel.save(filename='output.xlsx')


